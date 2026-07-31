import os
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from prospecting_engine import generate_prospects

BASE_DIR = Path(__file__).parent.resolve()
RESULTS_DIR = BASE_DIR / "results"

def build_master_dataset():
    prospects = generate_prospects()
    df_prospects = pd.DataFrame(prospects)
    
    # Standardize column order
    cols = [
        "target_domain",
        "company_name",
        "company_website",
        "industry",
        "contact_person",
        "job_title",
        "business_email",
        "personal_email",
        "contact_linkedin",
        "company_linkedin",
        "phone_number",
        "hq_location",
        "country",
        "provenance_repo",
        "source_url",
        "confidence_score",
        "prospect_use_case"
    ]
    
    df_master = df_prospects[cols].drop_duplicates(subset=["target_domain", "company_name", "business_email"])
    df_master = df_master.sort_values(by=["target_domain", "confidence_score"], ascending=[True, False])
    
    # Save CSV
    csv_path = BASE_DIR / "master_prospects.csv"
    df_master.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"Saved master_prospects.csv ({len(df_master)} rows)")
    
    # Save XLSX with formatting
    xlsx_path = BASE_DIR / "master_prospects.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master Prospects"
    
    # Header styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Navy
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    headers = [c.replace("_", " ").title() for c in cols]
    ws.append(headers)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Write data rows
    for row in df_master.itertuples(index=False):
        ws.append(list(row))
        
    for row_num in range(2, ws.max_row + 1):
        for col_num in range(1, len(cols) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=10)
            if col_num in [16]: # Confidence score
                cell.alignment = Alignment(horizontal="center")
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
        
    wb.save(xlsx_path)
    print(f"Saved master_prospects.xlsx formatted successfully.")

if __name__ == "__main__":
    build_master_dataset()
